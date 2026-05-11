import openpyxl
from openpyxl.styles import Border, Font, Side



class SheetsExporter:
    def __init__(self, columns: list[str], max_scores: dict[str, float | int] | None = None):
        self.columns: list[str] = columns
        self.sheet_items: dict[str, dict[str,float|None]] = {}
        self.max_scores: dict[str, float | int] | None = max_scores
        self.student_names: dict[str, str] = {}

    def add_student(self, key_name: str, student_name: str):
        self.sheet_items[key_name] = {i: -1 for i in self.columns}
        self.student_names[key_name] = student_name

    def append(self, key_name: str, num_to_score: dict[str, float|None]):
        if key_name not in self.sheet_items:
            print("WARNING: Student not in sheet; append operation aborted.")
            return
        self.sheet_items[key_name] = num_to_score

    def export_sheet(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()

        sheet = wb.active
        assert sheet is not None
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 25

        total_col = len(self.columns) + 3

        for item in self.columns:
            sheet.cell(
                    row=1,
                    column=self.columns.index(item)+3,
                    value=item
                    )
        sheet.cell(row=1, column=total_col, value="Total")

        medium_bottom      = Border(bottom=Side(style="medium"))
        medium_right       = Border(right=Side(style="medium"))
        medium_both        = Border(bottom=Side(style="medium"), right=Side(style="medium"))
        medium_left        = Border(left=Side(style="medium"))
        medium_left_bottom = Border(left=Side(style="medium"), bottom=Side(style="medium"))

        sheet.cell(row=1, column=2).border = medium_right
        sheet.cell(row=1, column=total_col).border = medium_left

        data_start_row = 2
        if self.max_scores is not None:
            sheet.cell(row=2, column=1).border = medium_bottom
            sheet.cell(row=2, column=2).border = medium_both
            for item in self.columns:
                col = self.columns.index(item)+3
                cell = sheet.cell(row=2, column=col, value=self.max_scores.get(item, ""))
                cell.border = medium_bottom
            max_total = sum(self.max_scores.values())
            _fmt = lambda v: int(v) if isinstance(v, float) and v % 1 == 0 else v
            sheet.cell(row=2, column=total_col, value=f"{_fmt(max_total)}/{_fmt(max_total)}").border = medium_left_bottom
            data_start_row = 3

        students_list = list(self.sheet_items.keys())

        for idx, key_name in enumerate(students_list):
            row = idx+data_start_row
            sheet.cell(row=row, column=1, value=key_name)
            cell = sheet.cell(row=row, column=2, value=self.student_names[key_name])
            cell.border = medium_right

        for idx, key_name in enumerate(students_list):
            for item in self.columns:
                value = self.sheet_items[key_name][item]
                sheet.cell(
                        row=idx+data_start_row,
                        column=self.columns.index(item)+3,
                        value=value if value is not None else ""
                        )
            scores = self.sheet_items[key_name]
            student_total = sum(v for v in scores.values() if v is not None and v >= 0)
            if self.max_scores is not None:
                answered_items = [
                    item for item in self.columns
                    if (score_item := scores.get(item)) is not None
                    and score_item >= 0
                    ]
                answered_max = sum(self.max_scores[item] for item in answered_items if item in self.max_scores)
                _fmt = lambda v: int(v) if isinstance(v, float) and v % 1 == 0 else v
                sheet.cell(
                        row=idx+data_start_row,
                        column=total_col,
                        value=f"{_fmt(student_total)}/{_fmt(answered_max)}"
                        ).border = medium_left

        bold = Font(bold=True)
        for col in range(1, total_col + 1):
            sheet.cell(row=1, column=col).font = bold
        total_rows = data_start_row + len(students_list)
        for row in range(1, total_rows):
            sheet.cell(row=row, column=1).font = bold

        summary_start_row = data_start_row + len(students_list) + 3
        self._add_summary(sheet, summary_start_row)

        return wb

    def _add_summary(self, sheet, start_row: int):
        if self.max_scores is None:
            return

        student_pcts = []
        for scores in self.sheet_items.values():
            answered = [
                (item, v) for item, v in scores.items()
                if v is not None and v >= 0 and item in self.max_scores
            ]
            answered_max = sum(self.max_scores[item] for item, _ in answered)
            if answered_max <= 0:
                continue
            total = sum(v for _, v in answered)
            student_pcts.append(total / answered_max)
        class_mean_pct = (sum(student_pcts) / len(student_pcts) * 100) if student_pcts else 0.0

        item_pcts: dict[str, float] = {}
        for item in self.columns:
            if item not in self.max_scores or self.max_scores[item] <= 0:
                continue
            vals: list[float] = [
                s for scores in self.sheet_items.values()
                if (s := scores.get(item)) is not None and s >= 0
            ]
            if not vals:
                continue
            item_pcts[item] = (sum(vals) / len(vals)) / self.max_scores[item] * 100

        bold = Font(bold=True)
        sheet.cell(row=start_row, column=1, value="Summary").font = bold
        sheet.cell(row=start_row + 1, column=1, value="Class Mean").font = bold
        sheet.cell(row=start_row + 1, column=2, value=f"{class_mean_pct:.1f}%")

        if item_pcts:
            hi = max(item_pcts, key=lambda k: item_pcts[k])
            lo = min(item_pcts, key=lambda k: item_pcts[k])
            sheet.cell(row=start_row + 2, column=1, value="Highest Item").font = bold
            sheet.cell(row=start_row + 2, column=2, value=f"{hi} ({item_pcts[hi]:.1f}%)")
            sheet.cell(row=start_row + 3, column=1, value="Lowest Item").font = bold
            sheet.cell(row=start_row + 3, column=2, value=f"{lo} ({item_pcts[lo]:.1f}%)")
