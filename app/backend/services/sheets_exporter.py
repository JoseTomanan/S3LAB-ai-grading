import openpyxl



class SheetsExporter:
    def __init__(self, columns: list[str]):
        self.columns: list[str] = columns
        self.sheet_items: dict[str, dict[str,float|None]] = {}

    def add_student(self, key_name: str):
        self.sheet_items[key_name] = {i: -1 for i in self.columns}

    def append(self, key_name: str, num_to_score: dict[str, float|None]):
        if key_name not in self.sheet_items:
            print("WARNING: Student not in sheet; append operation aborted.")
            return
        self.sheet_items[key_name] = num_to_score

    def export_sheet(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        
        sheet = wb.active
        assert sheet is not None
        sheet.column_dimensions['A'].width = 25

        students_list = list(self.sheet_items.keys())

        for key_name in students_list:
            sheet.cell(
                    row=students_list.index(key_name)+2,
                    column=1,
                    value=key_name
                    )
        
        for item in self.columns:
            sheet.cell(
                    row=1,
                    column=self.columns.index(item)+2,
                    value=item
                    )

        for key_name in students_list:
            for item in self.columns:
                value = self.sheet_items[key_name][item]
                sheet.cell(
                        row=students_list.index(key_name)+2,
                        column=self.columns.index(item)+2,
                        value=value if value is not None else ""
                        )
        
        # sheets_directory = f"./static/sheets/{file_name}.xlsx"

        # wb.save(sheets_directory)
        # print(f"--> File '{file_name}.xlsx' saved.")

        return wb