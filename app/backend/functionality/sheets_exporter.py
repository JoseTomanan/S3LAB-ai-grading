import openpyxl



class SheetsExporter:
    def __init__(self, columns: list[str]):
        self.columns: list[str] = columns
        self.sheet_items: dict[str, dict[str,float]] = {}

    def add_student(self, key_name: str):
        self.sheet_items[key_name] = {i: -1 for i in self.columns}

    def append(self, key_name: str, num_to_score: dict[str, float]):
        if key_name not in self.sheet_items:
            print("WARNING: Student not in sheet; append operation aborted.")
            return
        self.sheet_items[key_name] = num_to_score

    def export_sheet(self, file_name: str):
        wb = openpyxl.Workbook()
        
        sheet = wb.active
        assert sheet is not None

        students_list = list(self.sheet_items.keys())

        for key_name in students_list:
            sheet.cell(
                    row=students_list.index(key_name)+2,
                    column=1,
                    value=key_name
                    )
        
        for item in self.columns:
            sheet.cell(
                    row=1,
                    column=self.columns.index(item)+2,
                    value=item
                    )

        for key_name in students_list:
            for item in self.columns:
                sheet.cell(
                        row=students_list.index(key_name)+2,
                        column=self.columns.index(item)+2,
                        value=self.sheet_items[key_name][item]
                        )
        
        sheets_directory = f"./temp_sheets_output/{file_name}.xlsx"

        wb.save(sheets_directory)
        print(f"--> File '{file_name}.xlsx' saved.")

        return sheets_directory